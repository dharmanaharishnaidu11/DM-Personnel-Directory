"""
Extract officer/contact records from 4 local Excel files.
Each file has a different structure - handled with custom parsers.
Output: local_excel_contacts.json
"""
import json
import os
import re
import traceback
from openpyxl import load_workbook

OUTPUT = r"d:\APSDMA_MASTER_DATA\07_DATABASES\09_GO_Doccuments\district_key_posts_export\local_excel_contacts.json"

def cs(v):
    """Clean cell to string."""
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() in ("nan", "none", "nat", "-", "__", "no"):
        return ""
    return s

def cp(v):
    """Clean phone - keep digits and commas for multi-numbers."""
    s = cs(v)
    if not s:
        return ""
    # Handle float artifacts
    if '.' in s:
        try:
            n = float(s)
            s = str(int(n))
        except:
            pass
    # Remove everything except digits, commas, spaces, newlines, +
    # Split on newline or comma to handle multiple numbers
    parts = re.split(r'[\n,/]+', s)
    cleaned = []
    for p in parts:
        digits = re.sub(r'[^\d]', '', p.strip())
        if len(digits) >= 7:
            cleaned.append(digits)
    return ", ".join(cleaned)

def rec(name, desig, district, phone, email, source, sheet=""):
    """Create a record dict, skip if no useful data."""
    name = cs(name)
    desig = cs(desig)
    district = cs(district)
    phone = cp(phone)
    email = cs(email)
    if not name and not phone:
        return None
    # Skip serial-number-only names
    if name and name.isdigit():
        name = ""
    if not name and not phone:
        return None
    return {
        "person_name": name,
        "designation": desig,
        "district": district,
        "phone": phone,
        "email": email,
        "source": source,
        "sheet": sheet,
    }


# =============================================================================
# FILE 1: 26 Districts contacts.xlsx
# Structure: 26 sheets (one per district). Each sheet:
#   Row 0: "Contact details"
#   Row 1: "Name of the District: XXX District"
#   Row 2: S.No | Rank | Name | Contact Number | Email ID
#   Row 3+: data
# =============================================================================
def extract_file1(filepath):
    fname = os.path.basename(filepath)
    records = []
    wb = load_workbook(filepath, read_only=True, data_only=True)

    for sname in wb.sheetnames:
        ws = wb[sname]
        district = sname.strip()
        sheet_recs = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            vals = list(row) if row else []
            if i < 3:  # skip header rows
                continue
            if len(vals) < 4:
                continue
            # cols: 0=S.No, 1=Rank/Designation, 2=Name, 3=Phone, 4=Email
            name = cs(vals[2]) if len(vals) > 2 else ""
            desig = cs(vals[1]) if len(vals) > 1 else ""
            phone = vals[3] if len(vals) > 3 else ""
            email = cs(vals[4]) if len(vals) > 4 else ""

            r = rec(name, desig, district, phone, email, fname, sname)
            if r:
                records.append(r)
                sheet_recs += 1

        print(f"  Sheet '{sname}': {sheet_recs} records")

    wb.close()
    return records


# =============================================================================
# FILE 2: All District contacts.xlsx
# Structure: Sheet 'Collectors' has wide layout:
#   Row 0: title
#   Row 1: Sl.No | District | Collector | JC | DRO | DPM | AO | D section
#   Row 2+: data - phone numbers with CC numbers embedded
# =============================================================================
def extract_file2(filepath):
    fname = os.path.basename(filepath)
    records = []
    wb = load_workbook(filepath, read_only=True, data_only=True)

    for sname in wb.sheetnames:
        ws = wb[sname]
        all_rows = list(ws.iter_rows(values_only=True))
        print(f"  Sheet '{sname}': {len(all_rows)} rows")

        if sname == 'Collectors':
            # Row 1 is header: Sl.No | District | Collector | JC | DRO | DPM | AO | D section
            # Data starts at row 2
            # Each cell has phone numbers (sometimes with "CC-xxx" for confidential clerk)
            designations = ['Collector', 'Joint Collector', 'DRO', 'DPM', 'AO', 'D Section']
            col_indices = [2, 3, 4, 5, 6, 7]  # columns for each designation

            for row_idx in range(2, len(all_rows)):
                row = all_rows[row_idx]
                if not row or len(row) < 3:
                    continue
                district = cs(row[1])
                if not district:
                    continue

                for desig_name, col_i in zip(designations, col_indices):
                    if col_i < len(row):
                        cell = cs(row[col_i])
                        if not cell:
                            continue
                        # Parse "main_number CC-cc_number" pattern
                        # Split into main officer phone and CC phone
                        parts = re.split(r'\s+CC[-\s]*', cell, flags=re.IGNORECASE)
                        main_phone = cp(parts[0]) if parts else ""
                        cc_phone = cp(parts[1]) if len(parts) > 1 else ""

                        if main_phone:
                            r = rec("", desig_name, district, main_phone, "", fname, sname)
                            if r:
                                records.append(r)
                        if cc_phone:
                            r = rec("", f"CC to {desig_name}", district, cc_phone, "", fname, sname)
                            if r:
                                records.append(r)

            print(f"    -> {len(records)} records from Collectors sheet")

    wb.close()
    return records


# =============================================================================
# FILE 3: Collecters contact list.xlsx
# Structure: Sheet1:
#   Row 0: title
#   Row 1: SI.No | District | Name | Designation | Contact Number | Fax | Email
#   Row 2+: data (district may be blank for continuation rows)
# =============================================================================
def extract_file3(filepath):
    fname = os.path.basename(filepath)
    records = []
    wb = load_workbook(filepath, read_only=True, data_only=True)

    ws = wb['Sheet1']
    all_rows = list(ws.iter_rows(values_only=True))
    current_district = ""

    for row_idx in range(2, len(all_rows)):
        row = all_rows[row_idx]
        if not row or len(row) < 5:
            continue
        # Update district if present
        d = cs(row[1])
        if d:
            current_district = d
        name = cs(row[2])
        desig = cs(row[3])
        phone = row[4]
        email = cs(row[6]) if len(row) > 6 else ""

        r = rec(name, desig, current_district, phone, email, fname, "Sheet1")
        if r:
            records.append(r)

    wb.close()
    return records


# =============================================================================
# FILE 4: ALL Collectors, JCs, DRO, RTGS, APSDMA, Agriculture Contact Numbers.xlsx
# Multiple sheets with different structures:
#   - Sheet1: wide layout with Collector/JC1/JC2/Agri columns (name+phone pairs)
#   - District Admin: similar wide layout with Collector/JC1/JC2/DRO/Agri
#   - with names: similar + APSDMA Staff column
#   - without names: phones only (skip - less useful)
#   - Final: similar to 'with names' but APSDMA first (skip - duplicate)
#   - From RTGS Final: simple name+phone list
#   - Sheet2: simple name+phone list
# =============================================================================
def extract_file4(filepath):
    fname = os.path.basename(filepath)
    records = []
    wb = load_workbook(filepath, read_only=True, data_only=True)
    all_sheets = {s: list(wb[s].iter_rows(values_only=True)) for s in wb.sheetnames}

    # --- Best structured sheet: "District Admin" ---
    # Row 0: headers (merged)
    # Row 1: sub-headers: [Sl.No, District, CollectorName, CollectorPhone, JC1Name, JC1Phone, JC2Name, JC2Phone, DROName, DROPhone, AgriName, AgriPhone]
    # Row 2+: data
    seen_combos = set()  # track (name, phone) to avoid duplicates across sheets

    rows = all_sheets.get('District Admin', [])
    for row_idx in range(2, len(rows)):
        row = rows[row_idx]
        if not row or len(row) < 4:
            continue
        district = cs(row[1])
        if not district:
            continue

        officer_pairs = [
            (2, 3, "District Collector"),
            (4, 5, "Joint Collector 1"),
            (6, 7, "Joint Collector 2"),
            (8, 9, "DRO"),
            (10, 11, "Agri Joint Director"),
        ]
        for name_col, phone_col, desig_name in officer_pairs:
            if name_col < len(row) and phone_col < len(row):
                name = cs(row[name_col])
                phone = row[phone_col]
                if name or cp(phone):
                    key = (name, cp(phone))
                    if key not in seen_combos:
                        seen_combos.add(key)
                        r = rec(name, desig_name, district, phone, "", fname, "District Admin")
                        if r:
                            records.append(r)

    print(f"  District Admin: {len(records)} records")

    # --- 'with names' sheet adds APSDMA Staff column ---
    rows = all_sheets.get('with names', [])
    apsdma_count = 0
    for row_idx in range(2, len(rows)):
        row = rows[row_idx]
        if not row or len(row) < 16:
            continue
        district = cs(row[1])
        # APSDMA staff: cols 14=name, 15=phone
        name = cs(row[14])
        phone = row[15]
        if name or cp(phone):
            key = (name, cp(phone))
            if key not in seen_combos:
                seen_combos.add(key)
                r = rec(name, "APSDMA Staff", district, phone, "", fname, "with names")
                if r:
                    records.append(r)
                    apsdma_count += 1

    print(f"  with names (APSDMA Staff): {apsdma_count} records")

    # --- From RTGS Final: simple list ---
    rows = all_sheets.get('From RTGS Final', [])
    rtgs_count = 0
    for row_idx in range(2, len(rows)):
        row = rows[row_idx]
        if not row or len(row) < 3:
            continue
        name = cs(row[1])
        phone = row[2]
        if name or cp(phone):
            key = (name, cp(phone))
            if key not in seen_combos:
                seen_combos.add(key)
                r = rec(name, "", "", phone, "", fname, "From RTGS Final")
                if r:
                    records.append(r)
                    rtgs_count += 1

    print(f"  From RTGS Final: {rtgs_count} records")

    # --- Sheet2: simple name+phone ---
    rows = all_sheets.get('Sheet2', [])
    s2_count = 0
    for row_idx in range(1, len(rows)):  # row 0 is header
        row = rows[row_idx]
        if not row or len(row) < 2:
            continue
        name = cs(row[0])
        phone = row[1]
        if name or cp(phone):
            key = (name, cp(phone))
            if key not in seen_combos:
                seen_combos.add(key)
                r = rec(name, "", "", phone, "", fname, "Sheet2")
                if r:
                    records.append(r)
                    s2_count += 1

    print(f"  Sheet2: {s2_count} records")

    wb.close()
    return records


def main():
    all_records = []
    file_counts = {}

    files = [
        (r"d:\APSDMA_MASTER_DATA\00_COMMAND_CENTER\SEOC_Dashboard\26 Districts contacts.xlsx", extract_file1),
        (r"d:\APSDMA_MASTER_DATA\01_ADMINISTRATION\04_Office_Admin\Existing\Jan2025_Docs\10012025\All District contacts.xlsx", extract_file2),
        (r"d:\APSDMA_MASTER_DATA\01_ADMINISTRATION\04_Office_Admin\Existing\ed1\All downloads\Collecters contact list.xlsx", extract_file3),
        (r"d:\APSDMA_MASTER_DATA\01_ADMINISTRATION\04_Office_Admin\Existing\ed1\All downloads\ALL Collectors, JCs, DRO, RTGS, APSDMA, Agriculture Contact Numbers.xlsx", extract_file4),
    ]

    for filepath, extractor in files:
        fname = os.path.basename(filepath)
        print(f"\n{'='*70}")
        print(f"Extracting: {fname}")
        print(f"{'='*70}")
        try:
            recs = extractor(filepath)
            file_counts[fname] = len(recs)
            all_records.extend(recs)
            print(f"  TOTAL: {len(recs)} records")
        except Exception as e:
            print(f"  ERROR: {e}")
            traceback.print_exc()
            file_counts[fname] = 0

    # Save
    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    with open(OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(all_records, f, indent=2, ensure_ascii=False)

    # Summary
    print(f"\n{'='*70}")
    print("SUMMARY")
    print(f"{'='*70}")
    for fname, count in file_counts.items():
        print(f"  {fname}: {count} records")
    print(f"  {'GRAND TOTAL':.<55} {len(all_records)}")
    print(f"\nSaved to: {OUTPUT}")

    # Sample records
    print("\nSample records:")
    for r in all_records[:8]:
        print(f"  {json.dumps(r, ensure_ascii=False)}")

    # Stats
    with_name = sum(1 for r in all_records if r['person_name'])
    with_phone = sum(1 for r in all_records if r['phone'])
    with_email = sum(1 for r in all_records if r['email'])
    with_desig = sum(1 for r in all_records if r['designation'])
    with_district = sum(1 for r in all_records if r['district'])
    print(f"\nField coverage:")
    print(f"  person_name: {with_name}/{len(all_records)}")
    print(f"  designation: {with_desig}/{len(all_records)}")
    print(f"  district:    {with_district}/{len(all_records)}")
    print(f"  phone:       {with_phone}/{len(all_records)}")
    print(f"  email:       {with_email}/{len(all_records)}")


if __name__ == '__main__':
    main()
